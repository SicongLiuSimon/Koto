"""
智能文件分析器 - 根据文件内容识别行业、类型、主题
支持 AI 增强分类（Ollama + Qwen3）
"""
import os
import json
import re
import requests
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from pathlib import Path
import mimetypes
import logging


logger = logging.getLogger(__name__)

class FileAnalyzer:
    """文件内容分析器（规则 + AI 混合分析）"""
    
    # AI 分类 Prompt（当规则匹配失败时使用本地 Ollama 模型）
    AI_CLASSIFY_PROMPT = """你是文件分类专家。根据文件名和内容摘要，判断文件所属类别。

可用类别（选一个最合适的）:
- finance: 金融/投资/合同/财务/收款/发票/协议
- startup: 创业公司BP/商业计划书/融资材料/公司介绍
- semiconductor: 半导体/芯片/HBM/晶圆/封装/电子元件
- academic: 学术论文/研究报告/实验/期刊/出版声明
- technology: AI/软件/编程/技术方案/演示/prototype
- career: 简历/招聘/求职/社会招聘/人事
- media: 电影/视频/影像/传媒/文化/历史/艺术
- medical: 医疗/医药/医院/健康/生物科技
- education: 教育/课程/培训/教材
- projects: 项目管理/进度/任务/需求/规划
- property: 物业/房产/租赁
- other: 无法归类

同时提取文件核心主题实体（如公司名、项目名、论文主题）。

只输出 JSON，不要任何解释：
{"industry": "类别", "category": "子类型(如contract/paper/bp/resume/report/presentation/software)", "entity": "核心实体名(简短)", "confidence": 0.0-1.0}"""

    # Ollama 连接配置
    OLLAMA_URL = "http://localhost:11434"
    AI_MODEL = "qwen3:8b"
    _ai_available = None
    _ai_check_time = 0
    
    def __init__(self):
        """初始化分类规则"""
        self.rules = self._load_classification_rules()
        
        # 行业中文标签（用于报告显示）
        self.industry_labels = {
            "finance": "金融投资",
            "startup": "创业公司",
            "semiconductor": "半导体",
            "academic": "学术研究",
            "technology": "技术/AI",
            "career": "求职招聘",
            "media": "传媒文化",
            "medical": "医疗健康",
            "education": "教育培训",
            "projects": "项目管理",
            "property": "物业房产",
            "other": "其他",
        }
        self.keywords_cache = {}
    
    def _load_classification_rules(self) -> Dict:
        """加载分类规则"""
        return {
            # 金融行业规则
            "finance": {
                "keywords": ["合同", "协议", "凭证", "发票", "财报", "融资", "投资", "预算", "账户", "资金"],
                "file_patterns": [r"合同", r"协议", r"发票", r"财报", r"预算"],
                "subcategories": {
                    "contract": ["合同", "协议", "协议书"],
                    "report": ["财报", "报告", "总结", "分析"],
                    "voucher": ["凭证", "发票", "收据"],
                    "investment": ["融资", "投资", "项目"],
                    "budget": ["预算", "费用", "支出"]
                }
            },
            # 房产行业规则
            "property": {
                "keywords": ["物业", "租赁", "房产", "维修", "缴费", "产权", "装修", "住户"],
                "file_patterns": [r"物业", r"租赁", r"房产", r"维修"],
                "subcategories": {
                    "property_management": ["物业", "管理", "服务"],
                    "rental": ["租赁", "租金", "合同"],
                    "maintenance": ["维修", "维护", "保养"],
                    "payment": ["缴费", "收款", "费用"]
                }
            },
            # 医疗行业规则
            "medical": {
                "keywords": ["病历", "处方", "诊断", "检查", "患者", "医疗", "健康", "费用"],
                "file_patterns": [r"病历", r"处方", r"诊断", r"检查"],
                "subcategories": {
                    "medical_record": ["病历", "记录", "档案"],
                    "prescription": ["处方", "药物", "治疗"],
                    "test": ["检查", "化验", "诊断"],
                    "billing": ["费用", "账单", "收费"]
                }
            },
            # 教育行业规则
            "education": {
                "keywords": ["课程", "作业", "成绩", "班级", "学生", "考试", "教材", "证书"],
                "file_patterns": [r"课程", r"作业", r"成绩", r"考试"],
                "subcategories": {
                    "course": ["课程", "教学", "讲座"],
                    "assignment": ["作业", "练习", "任务"],
                    "grade": ["成绩", "评分", "考试"],
                    "material": ["教材", "资料", "讲义"]
                }
            },
            # 项目管理规则
            "projects": {
                "keywords": ["项目", "任务", "计划", "进度", "里程碑", "团队", "需求"],
                "file_patterns": [r"项目", r"计划", r"进度"],
                "subcategories": {
                    "plan": ["计划", "规划", "方案"],
                    "progress": ["进度", "报告", "更新"],
                    "specification": ["需求", "规格", "说明"]
                }
            }
        }
    
    def analyze_file(self, file_path: str) -> Dict:
        """
        分析单个文件
        
        返回：
        {
            "file_name": str,
            "file_type": str,
            "industry": str,              # 行业分类
            "category": str,              # 文件类别
            "confidence": float,          # 置信度 0-1
            "keywords": List[str],        # 提取的关键词
            "suggested_folder": str,      # 建议的文件夹路径
            "timestamp": str,             # 提取的时间信息
            "metadata": Dict              # 扩展元数据
        }
        """
        file_path_obj = Path(file_path)
        
        if not file_path_obj.exists():
            return {
                "success": False,
                "error": f"文件不存在: {file_path}"
            }
        
        # 1. 提取文件基本信息
        file_name = file_path_obj.name
        file_type = file_path_obj.suffix.lower()
        file_size = file_path_obj.stat().st_size
        
        # 2. 提取文件内容和关键词
        content = self._extract_content(file_path)
        keywords = self._extract_keywords(file_name, content)
        
        # 3. 识别行业和类别
        industry, confidence = self._classify_industry(keywords, file_name, content)
        category = self._classify_category(industry, keywords)

        # 3.5 识别公司/项目实体
        entity_name, entity_type = self._extract_primary_entity(file_name, content)
        
        # ★ AI 增强：只要提取到文件内容就用本地 Ollama 模型分类（不再受置信度门槛限制）
        # 本地模型比关键词规则更准确；规则结果仅作兜底。
        ai_used = False
        ai_result = self._ai_classify(file_name, content, file_type) if content else None
        if ai_result:
            industry = ai_result.get("industry", industry)
            category = ai_result.get("category", category)
            confidence = ai_result.get("confidence", confidence)
            ai_entity = ai_result.get("entity")
            if ai_entity and not self._is_generic_name(ai_entity):
                entity_name = ai_entity
                entity_type = "ai_extracted"
            ai_used = True
            # ── 保存训练样本（异步，不阻塞主流程）────────────────────────
            self._save_training_sample(file_name, file_type, content, ai_result)
        elif confidence < 0.3:
            # 内容为空时才走原来的低置信兜底
            ai_result2 = self._ai_classify(file_name, content, file_type)
            if ai_result2:
                industry = ai_result2.get("industry", industry)
                category = ai_result2.get("category", category)
                confidence = ai_result2.get("confidence", confidence)
                ai_entity2 = ai_result2.get("entity")
                if ai_entity2 and not self._is_generic_name(ai_entity2):
                    entity_name = ai_entity2
                    entity_type = "ai_extracted"
                ai_used = True
        
        # 4. 提取时间信息
        timestamp = self._extract_timestamp(file_name, content)
        
        # 5. 生成建议文件夹路径
        suggested_folder = self._generate_folder_path(
            industry,
            category,
            timestamp,
            keywords,
            entity_name
        )
        
        return {
            "success": True,
            "file_name": file_name,
            "file_path": str(file_path),
            "file_type": file_type,
            "file_size": file_size,
            "industry": industry,
            "category": category,
            "confidence": confidence,
            "keywords": keywords,
            "timestamp": timestamp,
            "entity": entity_name,
            "entity_type": entity_type,
            "suggested_folder": suggested_folder,
            "ai_enhanced": ai_used,
            "preview": content[:500] if content else "(无法提取内容)"
        }
    
    def _ai_classify(self, file_name: str, content: str, file_type: str) -> Optional[Dict]:
        """使用 AI 模型进行智能文件分类
        
        本地模式：使用 Ollama + Qwen3（更快、免费）
        云端模式：使用 Gemini API（无需本地 GPU）
        当规则引擎置信度低时自动调用，利用 AI 理解文件语义。
        """
        import time
        import os
        
        is_cloud = os.environ.get('KOTO_DEPLOY_MODE') == 'cloud'
        
        # 构建文件摘要（给 AI 的上下文）
        content_preview = (content or "")[:800].strip()
        user_msg = f"文件名: {file_name}\n文件类型: {file_type}\n内容摘要: {content_preview[:500] if content_preview else '(无内容)'}"
        
        valid_industries = [
            "finance", "startup", "semiconductor", "academic", "technology",
            "career", "media", "medical", "education", "projects", "property", "other"
        ]
        
        raw = None
        
        if is_cloud:
            # ═══ 云端模式：使用 Gemini API ═══
            try:
                from google import genai as _genai
                from google.genai import types as _types
                
                api_key = os.environ.get('GEMINI_API_KEY') or os.environ.get('API_KEY')
                if not api_key:
                    return None
                
                _client = _genai.Client(api_key=api_key)
                resp = _client.models.generate_content(
                    model="gemini-2.0-flash-lite",
                    contents=f"{self.AI_CLASSIFY_PROMPT}\n\n{user_msg}",
                    config=_types.GenerateContentConfig(
                        temperature=0.0,
                        max_output_tokens=200,
                    )
                )
                raw = (resp.text or "").strip()
            except Exception as e:
                logger.error(f"[FileAnalyzer AI/Cloud] ❌ Gemini 分类失败: {e}")
                return None
        else:
            # ═══ 本地模式：使用 Ollama ═══
            # 检查 Ollama 可用性（缓存 60 秒）
            now = time.time()
            if self._ai_available is not None and (now - self._ai_check_time) < 60:
                if not self._ai_available:
                    return None
            
            try:
                import socket
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(0.2)
                result = sock.connect_ex(('127.0.0.1', 11434))
                sock.close()
                FileAnalyzer._ai_available = (result == 0)
                FileAnalyzer._ai_check_time = now
                if not FileAnalyzer._ai_available:
                    return None
            except Exception:
                FileAnalyzer._ai_available = False
                FileAnalyzer._ai_check_time = now
                return None
            
            try:
                resp = requests.post(
                    f"{self.OLLAMA_URL}/api/chat",
                    json={
                        "model": self.AI_MODEL,
                        "messages": [
                            {"role": "system", "content": self.AI_CLASSIFY_PROMPT},
                            {"role": "user", "content": user_msg},
                        ],
                        "stream": False,
                        "format": "json",
                        "think": False,  # Qwen3: 禁用思考，加速分类
                        "options": {
                            "temperature": 0.0,
                            "num_predict": 120,
                        }
                    },
                    timeout=8.0
                )
                
                if resp.status_code != 200:
                    return None
                
                raw = (resp.json().get("message", {}) or {}).get("content", "")
            except requests.exceptions.Timeout:
                logger.info(f"[FileAnalyzer AI] ⏱️ 超时: {file_name[:30]}")
                return None
            except Exception as e:
                logger.error(f"[FileAnalyzer AI] ❌ Ollama 错误: {e}")
                return None
        
        # ═══ 解析 AI 返回的 JSON ═══
        if not raw:
            return None
        
        try:
            # 尝试提取 JSON（Gemini 可能返回 markdown 代码块）
            import re as _re
            json_match = _re.search(r'\{[^{}]+\}', raw)
            if json_match:
                raw = json_match.group()
            
            data = json.loads(raw.strip())
            industry = str(data.get("industry", "")).strip().lower()
            category = str(data.get("category", "")).strip().lower()
            entity = str(data.get("entity", "")).strip()
            confidence = float(data.get("confidence", 0.0))
            
            # 验证 industry 合法
            if industry not in valid_industries:
                # 尝试模糊匹配
                for vi in valid_industries:
                    if vi in industry or industry in vi:
                        industry = vi
                        break
                else:
                    industry = "other"
            
            if confidence < 0.3:
                confidence = 0.6  # AI 给出结果视为至少 0.6 置信度
            
            source = "Cloud/Gemini" if is_cloud else "Local/Ollama"
            logger.info(f"[FileAnalyzer AI] ✅ {file_name[:30]} → {industry}/{category} ({confidence:.2f}) entity={entity} [{source}]")
            return {
                "industry": industry,
                "category": category,
                "entity": entity if entity else None,
                "confidence": confidence,
            }
        except Exception as e:
            logger.error(f"[FileAnalyzer AI] ❌ JSON 解析错误: {e}")
            return None

    def _extract_content(self, file_path: str) -> str:
        """提取文件文本内容"""
        file_path_obj = Path(file_path)
        file_type = file_path_obj.suffix.lower()
        
        # 文本文件直接读取
        if file_type in ['.txt', '.json', '.csv', '.log']:
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read(2000)  # 读取前2000字符
            except:
                return ""
        
        # PDF提取文本
        if file_type == '.pdf':
            try:
                import PyPDF2
                with open(file_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in reader.pages[:3]:  # 读前3页
                        text += page.extract_text()
                    return text[:2000]
            except:
                return ""
        
        # Office文档提取
        if file_type in ['.docx', '.doc']:
            try:
                from docx import Document
                doc = Document(file_path)
                text = "\n".join([p.text for p in doc.paragraphs[:20]])
                return text[:2000]
            except:
                return ""
        
        if file_type in ['.xlsx', '.xls']:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                text = ""
                for sheet in wb.sheetnames[:2]:
                    ws = wb[sheet]
                    for row in list(ws.iter_rows(values_only=True))[:10]:
                        text += " ".join(str(cell) for cell in row if cell)
                return text[:2000]
            except:
                return ""
        
        # 文件名作为备选内容
        return file_path_obj.name
    
    def _extract_keywords(self, file_name: str, content: str) -> List[str]:
        """从文件名和内容中提取关键词"""
        all_text = f"{file_name} {content}".lower()
        keywords = []
        
        # 合并所有规则的关键词并检测
        all_rule_keywords = []
        for industry_rules in self.rules.values():
            all_rule_keywords.extend(industry_rules.get("keywords", []))
        
        for keyword in set(all_rule_keywords):
            if keyword.lower() in all_text:
                keywords.append(keyword)
        
        return list(set(keywords))  # 去重

    def _is_generic_name(self, name: str) -> bool:
        """Heuristic filter for generic or low-signal names."""
        if not name:
            return True
        cleaned = re.sub(r"\s+", "", name.strip())
        if len(cleaned) < 2:
            return True
        generic_terms = {
            "报告", "总结", "访谈", "会议", "记录", "文档", "资料", "模板",
            "合同", "协议", "说明", "归纳", "计划", "项目", "方案", "需求",
            "分析", "手册", "说明书", "表格", "清单", "草案", "稿件",
            "访谈报告",
            "report", "summary", "interview", "document", "notes", "draft"
        }
        lower = cleaned.lower()
        if lower in generic_terms:
            return True
        if re.fullmatch(r"[0-9_\-]+", cleaned):
            return True
        return False

    def _save_training_sample(self, file_name: str, file_type: str, content: str, ai_result: dict) -> None:
        """将 AI 分类结果保存为本地训练样本 (JSONL)，供后续微调使用。
        同一文件名已存在则跳过（防止重复运行产生重复样本）。
        """
        try:
            import json
            from pathlib import Path as _P
            from datetime import datetime as _dt

            _train_dir = _P(__file__).parent.parent / "config" / "training_data"
            _train_dir.mkdir(parents=True, exist_ok=True)
            _train_file = _train_dir / "file_classify_samples.jsonl"

            # ── 去重：若该文件名已存在样本则跳过 ──────────────────────────
            if _train_file.exists():
                _seen: set = set()
                try:
                    with open(_train_file, "r", encoding="utf-8") as _rf:
                        for _line in _rf:
                            _obj = json.loads(_line)
                            # user 字段第一行 "文件名: <name>"
                            _u = _obj.get("user", "")
                            if _u.startswith("文件名: "):
                                _seen.add(_u.split("\n", 1)[0][len("文件名: "):])
                    if file_name in _seen:
                        return  # 已有此文件的分类样本，跳过
                except Exception:
                    pass  # 去重失败时仍继续写入，宁可重复也不丢数据

            content_preview = (content or "")[:600].strip()
            sample = {
                "system": (
                    "你是文件分类专家。根据文件名和内容摘要，输出 JSON 分类结果，"
                    "字段：industry, category, entity, confidence。"
                ),
                "user": f"文件名: {file_name}\n文件类型: {file_type}\n内容摘要: {content_preview}",
                "assistant": json.dumps(ai_result, ensure_ascii=False),
                "task_type": "FILE_CLASSIFY",
                "source": "catalog_run",
                "quality": float(ai_result.get("confidence", 0.7)),
                "timestamp": _dt.now().isoformat(),
            }
            with open(_train_file, "a", encoding="utf-8") as f:
                f.write(json.dumps(sample, ensure_ascii=False) + "\n")
        except Exception:
            pass  # 训练样本写入失败不影响主流程

    def _sanitize_component(self, value: str) -> str:
        """Clean a path component to avoid invalid characters."""
        if not value:
            return ""
        replacements = {
            "\\": "_",
            "/": "_",
            ":": "_",
            "*": "_",
            "?": "_",
            '"': "_",
            "<": "_",
            ">": "_",
            "|": "_"
        }
        for char, replacement in replacements.items():
            value = value.replace(char, replacement)
        return value.strip()

    # ========== 文件名清理：去除版本/修订/重复后缀 ==========

    _REVISION_PATTERNS = [
        r'_revised(?:\(\d+\))?$',       # _revised, _revised(1)
        r'\(\d+\)$',                     # (1), (2)
        r'_\d{8}_\d{6}$',               # _20260203_004341
        r'_\d{14}$',                    # _20260203004341
        r'_copy$',                       # _copy
        r'\s*-\s*副本$',                 # - 副本
        r'\s*\(副本\)$',                 # (副本)
        r'_\d+$',                        # _1, _2 (trailing counter)
    ]

    def _clean_filename_stem(self, stem: str) -> str:
        """去除文件名中的版本号、修订后缀、重复下载后缀等，提取核心主题名。
        
        Examples:
            '电影时间的计算解析_revised(3)' -> '电影时间的计算解析'
            '数字之眼的危机_revised_20260215_024127' -> '数字之眼的危机'
            '原文_revised(1)' -> '原文'
            '论文出版声明（知网）-译者(1)' -> '论文出版声明（知网）-译者'
        """
        cleaned = stem
        # 反复应用直到稳定
        for _ in range(5):
            prev = cleaned
            for pattern in self._REVISION_PATTERNS:
                cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
            cleaned = cleaned.strip(' _-')
            if cleaned == prev:
                break
        return cleaned if cleaned else stem

    def _extract_primary_entity(self, file_name: str, content: str) -> Tuple[Optional[str], Optional[str]]:
        """Extract the most relevant company or project name from file name/content.
        
        优先级:
        1. 内容中标注的公司/项目名（公司名称：XXX）
        2. 内容中的公司名后缀（XX有限公司）
        3. 清理后的文件名主题（去除版本/修订后缀）
        4. 英文项目名
        """
        file_stem = Path(file_name).stem
        # 先清理文件名后缀
        cleaned_stem = self._clean_filename_stem(file_stem)

        # 1) 标注的实体
        labeled_patterns = [
            (r"公司名称[:：]\s*([^\n\r，,。;；]{2,30})", "company"),
            (r"企业名称[:：]\s*([^\n\r，,。;；]{2,30})", "company"),
            (r"机构名称[:：]\s*([^\n\r，,。;；]{2,30})", "company"),
            (r"品牌名称[:：]\s*([^\n\r，,。;；]{2,30})", "company"),
            (r"项目名称[:：]\s*([^\n\r，,。;；]{2,30})", "project")
        ]

        for pattern, entity_type in labeled_patterns:
            match = re.search(pattern, content)
            if match:
                name = self._sanitize_component(match.group(1))
                if not self._is_generic_name(name):
                    return name, entity_type

        # 2) 公司名后缀
        company_pattern = r"([\u4e00-\u9fff]{2,20}(?:公司|有限公司|集团|股份|医疗|科技|生物|医院|研究院|实验室|工作室))"
        match = re.search(company_pattern, content)
        if match:
            name = self._sanitize_component(match.group(1))
            if not self._is_generic_name(name):
                return name, "company"

        # 3) 清理后的文件名（已去除 _revised(N) 等后缀）
        if cleaned_stem and not self._is_generic_name(cleaned_stem):
            name = self._sanitize_component(cleaned_stem)
            if re.fullmatch(r"[A-Za-z][A-Za-z0-9&\-]{2,30}", name):
                return name, "project"
            return name, "document"

        # 4) 英文项目名
        english_match = re.search(r"\b([A-Z][A-Za-z0-9&\-]{2,30})\b", content)
        if english_match:
            name = self._sanitize_component(english_match.group(1))
            if not self._is_generic_name(name):
                return name, "project"

        return None, None
    
    def _classify_industry(self, keywords: List[str], file_name: str, content: str) -> Tuple[str, float]:
        """根据关键词分类行业"""
        industry_scores = {}
        
        for industry, rules in self.rules.items():
            score = 0
            total = 0
            
            # 关键词匹配
            for keyword in rules.get("keywords", []):
                total += 1
                if keyword in keywords:
                    score += 1
            
            # 文件名模式匹配
            for pattern in rules.get("file_patterns", []):
                if re.search(pattern, file_name, re.IGNORECASE):
                    score += 2
            
            if total > 0:
                industry_scores[industry] = score / (total + 1)
        
        if not industry_scores:
            return "other", 0.3
        
        best_industry = max(industry_scores.items(), key=lambda x: x[1])
        return best_industry[0], min(best_industry[1], 1.0)
    
    def _classify_category(self, industry: str, keywords: List[str]) -> str:
        """在行业内细分文件类别"""
        if industry not in self.rules:
            return "document"
        
        subcategories = self.rules[industry].get("subcategories", {})
        
        for category, category_keywords in subcategories.items():
            for keyword in category_keywords:
                if keyword in keywords or keyword in " ".join(keywords):
                    return category
        
        return "document"  # 默认分类
    
    def _extract_timestamp(self, file_name: str, content: str) -> Optional[str]:
        """从文件名和内容中提取时间信息"""
        all_text = f"{file_name} {content}"
        
        # 查找年份
        year_match = re.search(r'(20\d{2})', all_text)
        if year_match:
            year = year_match.group(1)
            
            # 查找月份
            month_match = re.search(rf'{year}[年\-/](\d{{1,2}})', all_text)
            if month_match:
                month = int(month_match.group(1))
                return f"{year}-{month:02d}"
            
            return year
        
        return None
    
    def _generate_folder_path(self, industry: str, category: str, timestamp: Optional[str], keywords: List[str], entity_name: Optional[str] = None) -> str:
        """生成建议的文件夹路径
        
        策略:
        1. 如果有实体名，归入 industry/entity_name 下（如 finance/华芯长晟）
        2. 如果实体名是通用名（'合同'、'报告'等），归入 industry/category
        3. 无实体名时，归入 industry/category
        """
        if entity_name:
            # 再次清理一下确保干净
            entity_name = self._clean_filename_stem(entity_name)
            entity_name = self._sanitize_component(entity_name)
            
            if entity_name and not self._is_generic_name(entity_name):
                # 用行业分类做顶层，实体名做子文件夹
                return f"{industry}/{entity_name}"

        # 无实体名或通用名 → 用行业+类别
        if category and category != "document":
            return f"{industry}/{category}"
        return industry
    
    def analyze_batch(self, file_paths: List[str]) -> List[Dict]:
        """批量分析文件"""
        results = []
        for file_path in file_paths:
            result = self.analyze_file(file_path)
            results.append(result)
        return results


# 快速测试
if __name__ == "__main__":
    analyzer = FileAnalyzer()
    
    # 示例文件分析
    test_files = [
        "sample_contract.txt",
        "financial_report.txt",
        "property_lease.txt"
    ]
    
    for test_file in test_files:
        if os.path.exists(test_file):
            result = analyzer.analyze_file(test_file)
            logger.info(f"\n分析结果: {result['file_name']}")
            logger.info(f"  行业: {result.get('industry', 'N/A')}")
            logger.info(f"  类别: {result.get('category', 'N/A')}")
            logger.info(f"  建议文件夹: {result.get('suggested_folder', 'N/A')}")
