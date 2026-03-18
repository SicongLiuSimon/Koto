#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据提取与转换模块 - 跨应用数据搬运专家
支持从微信、浏览器等提取数据并转换到Excel、数据库等
"""

import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class DataExtractor:
    """数据提取器基类"""

    def extract(self, source: Any) -> List[Dict[str, Any]]:
        """从数据源提取结构化数据"""
        raise NotImplementedError


class WeChatContactExtractor(DataExtractor):
    """微信联系人数据提取器"""

    def __init__(self):
        self.patterns = {
            "phone": r"1[3-9]\d{9}",  # 手机号
            "email": r"[\w\.-]+@[\w\.-]+\.\w+",  # 邮箱
            "wechat_id": r"[a-zA-Z][a-zA-Z0-9_-]{5,19}",  # 微信号
        }

    def extract_from_text(self, text: str) -> Dict[str, Any]:
        """从文本提取联系信息"""
        data = {"raw_text": text, "phones": [], "emails": [], "wechat_ids": []}

        # 提取手机号
        phones = re.findall(self.patterns["phone"], text)
        data["phones"] = list(set(phones))

        # 提取邮箱
        emails = re.findall(self.patterns["email"], text)
        data["emails"] = list(set(emails))

        return data

    def extract_from_wechat_chat(self, chat_history: List[str]) -> List[Dict[str, Any]]:
        """从微信聊天记录提取联系人信息"""
        contacts = []

        for message in chat_history:
            contact_data = self.extract_from_text(message)
            if contact_data["phones"] or contact_data["emails"]:
                contacts.append(contact_data)

        return contacts


class DataTransformer:
    """数据转换器 - 将提取的数据转换为目标格式"""

    @staticmethod
    def to_excel(
        data: List[Dict[str, Any]], output_path: str, sheet_name: str = "数据"
    ) -> str:
        """转换为Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            if not data:
                wb.save(output_path)
                return output_path

            # 获取所有列名
            all_keys = set()
            for item in data:
                all_keys.update(item.keys())
            headers = sorted(list(all_keys))

            # 设置标题样式
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 写入标题
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # 写入数据
            for row_idx, item in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    # 处理列表类型
                    if isinstance(value, (list, tuple)):
                        value = ", ".join(str(v) for v in value)
                    elif isinstance(value, dict):
                        value = json.dumps(value, ensure_ascii=False)

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # 保存
            wb.save(output_path)
            return output_path

        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

    @staticmethod
    def to_json(data: List[Dict[str, Any]], output_path: str, indent: int = 2) -> str:
        """转换为JSON文件"""
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)
        return output_path

    @staticmethod
    def to_csv(data: List[Dict[str, Any]], output_path: str) -> str:
        """转换为CSV文件"""
        import csv

        if not data:
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                pass
            return output_path

        # 获取所有列名
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())
        headers = sorted(list(all_keys))

        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()

            for item in data:
                row = {}
                for key in headers:
                    value = item.get(key, "")
                    if isinstance(value, (list, tuple)):
                        value = ", ".join(str(v) for v in value)
                    elif isinstance(value, dict):
                        value = json.dumps(value, ensure_ascii=False)
                    row[key] = value
                writer.writerow(row)

        return output_path


class CrossAppDataPipeline:
    """跨应用数据管道 - 场景1核心功能"""

    def __init__(self):
        self.extractors = {
            "wechat_contact": WeChatContactExtractor(),
        }
        self.transformer = DataTransformer()

    def run_pipeline(
        self, source_type: str, source_data: Any, target_format: str, output_path: str
    ) -> Dict[str, Any]:
        """
        执行数据管道

        Args:
            source_type: 数据源类型 (wechat_contact, browser_table, etc.)
            source_data: 源数据
            target_format: 目标格式 (excel, csv, json)
            output_path: 输出路径

        Returns:
            执行结果
        """
        try:
            # 1. 提取数据
            if source_type == "wechat_contact":
                extractor = self.extractors["wechat_contact"]
                if isinstance(source_data, str):
                    # 单个文本
                    extracted = [extractor.extract_from_text(source_data)]
                elif isinstance(source_data, list):
                    # 聊天记录列表
                    extracted = extractor.extract_from_wechat_chat(source_data)
                else:
                    return {"success": False, "error": "不支持的数据格式"}
            else:
                return {"success": False, "error": f"不支持的数据源类型: {source_type}"}

            # 2. 转换格式
            if target_format == "excel":
                result_path = self.transformer.to_excel(extracted, output_path)
            elif target_format == "csv":
                result_path = self.transformer.to_csv(extracted, output_path)
            elif target_format == "json":
                result_path = self.transformer.to_json(extracted, output_path)
            else:
                return {"success": False, "error": f"不支持的目标格式: {target_format}"}

            return {
                "success": True,
                "output_path": result_path,
                "record_count": len(extracted),
                "message": f"成功转换 {len(extracted)} 条数据到 {result_path}",
            }

        except Exception as e:
            return {"success": False, "error": str(e)}


# ================= 测试示例 =================

if __name__ == "__main__":
    # 测试数据提取
    sample_chat = [
        "张三的手机号是13812345678，邮箱是zhangsan@example.com",
        "李四：我的联系方式 15912345678",
        "王五的电话：13698745612，微信号：wangwu_wx",
    ]

    pipeline = CrossAppDataPipeline()

    # 测试转换到Excel
    result = pipeline.run_pipeline(
        source_type="wechat_contact",
        source_data=sample_chat,
        target_format="excel",
        output_path="workspace/documents/联系人信息.xlsx",
    )

    logger.info(json.dumps(result, ensure_ascii=False, indent=2))
